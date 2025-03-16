import logging
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from constants import LOGGER_NAME

logger = logging.getLogger(LOGGER_NAME)


class View:
    def __init__(self, repo, model):
        self.repo = repo
        self.model = model
        self.config = repo.config
        # streamlit parameters
        self.streamlit_settings = self.repo.config["streamlit"]["settings"]
        self.streamlit_widgets_config = self.repo.config["streamlit"]["widgets"]
        self.streamlit_captions = self.repo.config["streamlit"]["captions"]
        self.streamlit_widgets = dict()

        st.set_page_config(
            page_title=self.streamlit_settings["page_title"],
            layout=self.streamlit_settings["layout"],
            initial_sidebar_state=self.streamlit_settings["initial_sidebar_state"],
        )

    def show_title(self):
        st.title(self.streamlit_settings["title"])

    def show_run_button(self):
        return st.sidebar.button(self.streamlit_widgets_config["start_button"]["label"])

    def show_reset_button(self):
        return st.sidebar.button(self.streamlit_widgets_config["reset_button"]["label"])

    def show_select_names(self):
        return st.sidebar.multiselect(
            label=self.streamlit_widgets_config["select_names"]["label"],
            options=self.repo.names,
            default=self.repo.names,
        )

    def show_export_button(self):
        return st.sidebar.button(self.streamlit_widgets_config["export_button"]["label"])

    def show_success_message(self):
        st.success(self.streamlit_settings["success_message"])

    def show_toggle_data(self):
        return st.sidebar.toggle(
            self.streamlit_widgets_config["toggle_data"]["label"], value=False
        )

    def show_toggle_other_data(self):
        return st.sidebar.toggle(
            self.streamlit_widgets_config["toggle_other_data"]["label"], value=False
        )

    def show_multiselect_columns(self):
        columns = self.model.get_data_as_dataframe().columns
        return st.sidebar.multiselect(self.streamlit_widgets_config["multiselect_columns"]["label"], options=columns,
                                      default=columns)

    def show_slider_results(self):
        min_value = self.model.get_data_as_dataframe()[self.repo.column_infos["value"].col_name].min()
        max_value = self.model.get_data_as_dataframe()[self.repo.column_infos["value"].col_name].max()

        step = self.streamlit_widgets_config["slider_values"]['step']
        caption = self.streamlit_widgets_config["slider_values"]['label']
        return st.sidebar.slider(caption, min_value=min_value, max_value=max_value,
                                  value=max_value, step=step)

    def show_selected_items(self):
        values = self.model.get_data_as_dataframe()[self.repo.column_infos["item"].col_name].unique().tolist()
        return st.sidebar.pills(self.streamlit_widgets_config["pills_items"]["label"], values, selection_mode="multi",
                                default=values)

    def show_data(self):
        st.write(self.streamlit_captions["data"])
        st.dataframe(self.repo.get_data_as_dataframe())

    def show_other_data(self):
        st.write(self.streamlit_captions["other_data"])
        st.dataframe(self.repo.get_other_data_as_dataframe())

    def show_results(self, columns, items, limit):
        df = self.model.get_data_as_dataframe()
        logger.warning(df[self.repo.column_infos["name"].col_name].isin(items))
        df = df.loc[df[self.repo.column_infos["item"].col_name].isin(items)]
        df = df.loc[df[self.repo.column_infos["value"].col_name] <= limit, columns]
        logger.debug(f"Results | subset columns={columns}")
        st.write(self.streamlit_captions["results"])
        logger.debug(f"results shape={df.shape}")
        st.dataframe(df)

    def persist_value(self, key, value=True):
        st.session_state[key] = value

    def get_persisted_value(self, key):
        return st.session_state.get(key)

    def export_to_excel(self):
        workbooks = {}
        for name, path in self.config["insertions"]["excel_paths"].items():
            workbooks[name] = path

        for insertion, infos in self.config["insertions"]["mapping"].items():
            logger.info(f"insertion={insertion}")
            for name, xl_info in infos["excels"].items():
                logger.info(xl_info)
                wb = load_workbook(workbooks[name])
                logger.info(f"sheets={wb.sheetnames}")
                ws = wb[xl_info["sheet"]]
                val = getattr(getattr(self, infos["module"]), infos["method"])()
                if isinstance(val, pd.DataFrame):
                    logger.debug(f"insert self.{infos["module"]}.{infos["method"]} val={val.shape}")
                else:
                    logger.debug(f"insert self.{infos["module"]}.{infos["method"]} val={val}")
                # ws[xl_info["address"]].value = val
                # wb.save(workbooks[name])
